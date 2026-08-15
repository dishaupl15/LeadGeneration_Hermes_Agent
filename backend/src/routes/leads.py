"""
src/routes/leads.py
--------------------
Leads API router.

Production flow for POST /leads/generate-leads:
  UI → FastAPI → Google Maps discovery → CompanyEnrich → Serper → Firecrawl
  → NORMALIZE → VALIDATE → CONFIDENCE
  → [MONGODB PRE-DEDUP]  ← check existing leads by website/company_name
  → INSERT new leads only (skip existing)
  → return ONLY newly inserted leads to UI

Category-wise storage:
  Every generated lead is stored in a collection named leads_{category_slug}.
  e.g.  Construction → leads_construction
        Real Estate  → leads_real_estate
  The 'categories' collection lists all known industry names (seeded on startup).

Endpoints
---------
GET    /leads/categories        list all industry categories (from DB)
POST   /leads/generate-leads    pipeline → upsert MongoDB → return NEW leads only
GET    /leads                   paginated list of stored leads
POST   /leads                   create a lead manually
GET    /leads/{lead_id}         fetch a single lead by ID
PATCH  /leads/{lead_id}         partially update a lead
DELETE /leads/{lead_id}         permanently delete a lead
PATCH  /leads/{lead_id}/status  update CRM status for a lead
GET    /debug/database          MongoDB connectivity + document count
GET    /debug/sample            first 5 documents from the leads collection
"""

import importlib.util
import os
import time
import uuid as _uuid_mod
from datetime import datetime, timezone
from typing import Any, Optional
from urllib.parse import urlparse

from fastapi import APIRouter, HTTPException, Query, status

from src.config.mongo import (
    COLLECTION_NAME,
    CATEGORIES_COLLECTION,
    ALL_CATEGORIES,
    get_db,
    collection_for_category,
    ensure_lead_indexes,
)
from src.controllers.lead_controller import LeadController, LeadNotFoundError

HISTORY_COLLECTION = "generation_history"


def _make_run_id() -> str:
    """Generate a short unique run ID like RUN-a3f9b2c1."""
    return "RUN-" + _uuid_mod.uuid4().hex[:8].upper()


async def _create_history_run(db, run_id: str, category: str, query: str,
                               state: str, district: str, target: int,
                               filters: dict) -> None:
    """Insert a new generation run document with status=running."""
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "run_id": run_id,
        "category": category,
        "search_query": query,
        "state": state,
        "district": district,
        "requested_count": target,
        "generated_count": 0,
        "updated_count": 0,
        "status": "running",
        "started_at": now,
        "created_at": now,
        "completed_at": None,
        "failed_at": None,
        "duration_seconds": None,
        "source": "lead_generation",
        "filters": filters,
        "lead_ids": [],
        "logs": [
            {
                "timestamp": datetime.now().strftime("%H:%M:%S"),
                "level": "INFO",
                "stage": "init",
                "message": f"Generation run started — category={category!r} target={target}",
            }
        ],
        "statistics": {},
        "error_message": None,
        "pipeline_stats": None,
    }
    try:
        await db[HISTORY_COLLECTION].insert_one(doc)
    except Exception as exc:
        print(f"[HISTORY] WARNING: could not create run doc: {exc}", flush=True)


async def _append_log(db, run_id: str, level: str, stage: str, message: str) -> None:
    """Append one log entry to an existing run document."""
    entry = {
        "timestamp": datetime.now().strftime("%H:%M:%S"),
        "level": level,
        "stage": stage,
        "message": message,
    }
    try:
        await db[HISTORY_COLLECTION].update_one(
            {"run_id": run_id},
            {"$push": {"logs": entry}},
        )
    except Exception:
        pass  # non-fatal


async def _update_history_run(db, run_id: str, update: dict) -> None:
    """Apply a $set update to the history run document."""
    try:
        await db[HISTORY_COLLECTION].update_one(
            {"run_id": run_id},
            {"$set": update},
        )
    except Exception as exc:
        print(f"[HISTORY] WARNING: could not update run {run_id}: {exc}", flush=True)


async def _complete_history_run(db, run_id: str, inserted: int, updated: int,
                                 lead_ids: list[str], pipeline_stats: dict,
                                 stats: dict, t_start: float, logs_extra: list) -> None:
    """Mark a run as completed and store final statistics."""
    now = datetime.now(timezone.utc)
    duration = round(time.monotonic() - t_start, 1)
    for entry in logs_extra:
        await _append_log(db, run_id, entry["level"], entry["stage"], entry["message"])
    await _update_history_run(db, run_id, {
        "status": "completed",
        "generated_count": inserted,
        "updated_count": updated,
        "completed_at": now.isoformat(),
        "duration_seconds": duration,
        "lead_ids": lead_ids,
        "statistics": stats,
        "pipeline_stats": pipeline_stats,
    })


async def _fail_history_run(db, run_id: str, error_msg: str, t_start: float) -> None:
    """Mark a run as failed."""
    now = datetime.now(timezone.utc)
    duration = round(time.monotonic() - t_start, 1)
    await _append_log(db, run_id, "ERROR", "pipeline", f"Generation failed: {error_msg}")
    await _update_history_run(db, run_id, {
        "status": "failed",
        "failed_at": now.isoformat(),
        "duration_seconds": duration,
        "error_message": error_msg,
    })
from src.models.lead import Category
from src.schemas.lead_schema import (
    ErrorResponse,
    GenerateLeadsRequest,
    LeadCreateRequest,
    LeadResponse,
    LeadUpdateRequest,
    LeadsListResponse,
    MessageResponse,
    MongoLeadsResponse,
)

router = APIRouter(
    tags=["Leads"],
    responses={
        404: {"model": ErrorResponse, "description": "Lead not found"},
        422: {"description": "Validation error"},
    },
)


def _handle_not_found(exc: LeadNotFoundError) -> HTTPException:
    return HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))


def _ts() -> str:
    return datetime.now().strftime("%H:%M:%S")


def _log(tag: str, msg: str) -> None:
    print(f"[{_ts()}] [{tag}] {msg}", flush=True)


# ── Load leadgen pipeline module once at import time ─────────────────────────
_LEADGEN_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "..", "tools", "leadgen.py")
)
_leadgen_mod = None


def _get_leadgen():
    global _leadgen_mod
    if _leadgen_mod is not None:
        return _leadgen_mod
    if not os.path.exists(_LEADGEN_PATH):
        return None
    try:
        spec = importlib.util.spec_from_file_location("leadgen", _LEADGEN_PATH)
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _leadgen_mod = mod
        return mod
    except Exception as exc:
        _log("PIPELINE", f"WARNING: could not load leadgen.py: {exc}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# GET /leads/categories
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/leads/categories",
    summary="List all industry categories",
    response_model=list[str],
    tags=["Leads"],
)
async def get_categories():
    """
    Return all known industry category names.

    Reads from the MongoDB 'categories' collection which is seeded on startup
    with ALL_CATEGORIES from src/config/mongo.py.
    Falls back to the static list if MongoDB is unreachable.
    """
    try:
        db   = get_db()
        coll = db[CATEGORIES_COLLECTION]
        cursor = coll.find({}, {"name": 1, "_id": 0}).sort("name", 1)
        docs   = await cursor.to_list(length=500)
        names  = [d["name"] for d in docs if d.get("name")]
        if names:
            return names
    except Exception:
        pass
    return ALL_CATEGORIES


# ─────────────────────────────────────────────────────────────────────────────
# POST /leads/generate-leads
# ─────────────────────────────────────────────────────────────────────────────

@router.post(
    "/leads/generate-leads",
    summary="Generate B2B leads via Google Maps pipeline, store by category, return only NEW leads",
    response_model=MongoLeadsResponse,
    response_model_exclude_none=False,
    status_code=status.HTTP_200_OK,
    tags=["Leads"],
)
async def generate_leads(payload: GenerateLeadsRequest):
    """
    Production pipeline — Google Maps is the FIRST discovery layer.

    Stages
    ------
    1. [DISCOVERY]   Google Maps Places API → unique companies
    2. [ENRICH]      CompanyEnrich → Serper → Firecrawl (missing fields only)
    3. [CONFIDENCE]  Score 0.0-1.0 per company
    4. [ROUTE DEDUP] Remove within-batch duplicates by website domain
    5. [DB DEDUP]    Query MongoDB: find which candidates already exist
                     (by website URL or company_name as fallback)
    6. [MONGODB]     Upsert ALL into leads_{category} collection
                     (keeps data fresh even for existing leads)
    7. [RESPONSE]    Return ONLY the newly inserted leads to the UI

    Category storage
    ----------------
    Every lead is written to a collection named leads_{category_slug}:
      Construction  → leads_construction
      Real Estate   → leads_real_estate
      FinTech       → leads_fintech
    The 'categories' collection is updated with the industry name on every run.
    """
    t_start = time.monotonic()
    _log("LEADS", "Request received")

    from app.services.companyenrich_service import reset_credits_flag
    reset_credits_flag()

    # ── Resolve request params ────────────────────────────────────────────────
    import re as _re
    industry = payload.industry or ""
    state    = payload.state or ""
    district = payload.district or payload.city or ""
    target   = payload.resolved_target()

    if payload.query and not industry:
        query = payload.query.strip()
        m = _re.match(r'^(.+?)\s+companies?\s+in\s+(.+)$', query, _re.IGNORECASE)
        if m:
            industry = m.group(1).strip()
            if not state:
                district = m.group(2).strip()
    else:
        try:
            query = payload.resolved_query()
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=str(exc),
            )

    if not industry:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Provide 'industry' (and optionally 'state', 'district', 'target').",
        )
    if not state:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="'state' is required for lead generation (e.g. 'Maharashtra').",
        )

    query = f"{industry} companies in {district or state}, India"
    _log("LEADS", f"industry={industry!r} state={state!r} district={district!r} target={target}")

    # ── Create generation history run ─────────────────────────────────────────
    db_early = get_db()
    run_id = _make_run_id()
    await _create_history_run(
        db_early, run_id,
        category=industry,
        query=query,
        state=state,
        district=district,
        target=target,
        filters={"industry": industry, "state": state, "district": district, "target": target},
    )
    _log("HISTORY", f"Run created: {run_id}")

    # ── Stage 1: Google Maps pipeline ─────────────────────────────────────────
    _log("PIPELINE", "Starting Google Maps discovery")
    await _append_log(db_early, run_id, "SEARCH", "discovery", "Google Maps discovery started")
    from app.services.maps_pipeline_service import run_maps_pipeline, get_pipeline_stats
    try:
        maps_result = await run_maps_pipeline(
            category=industry,
            state=state,
            district=district or None,
            target=target,
            exclude_seen=True,
        )
    except Exception as exc:
        _log("DISCOVERY", f"Google Maps pipeline ERROR — {type(exc).__name__}: {exc}")
        await _fail_history_run(db_early, run_id, str(exc), t_start)
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(exc))

    pipeline_companies: list[dict] = maps_result.get("companies", [])
    ps = maps_result.get("pipeline_stats", {})
    _log("PIPELINE", f"Google Maps pipeline returned {len(pipeline_companies)} companies")
    await _append_log(db_early, run_id, "SEARCH", "discovery",
                      f"{ps.get('google_maps_discovered', len(pipeline_companies))} candidate companies discovered")

    if not pipeline_companies:
        _log("LEADS", "No companies found — returning empty result")
        await _append_log(db_early, run_id, "COMPLETE", "pipeline", "No companies found — run completed with 0 results")
        await _complete_history_run(db_early, run_id, 0, 0, [], {}, {}, t_start, [])
        return MongoLeadsResponse(
            success=True,
            inserted=0,
            updated=0,
            total=0,
            query=query,
            timestamp=datetime.now(timezone.utc).isoformat(),
            leads=[],
            pipeline_stats={
                "google_maps_discovered": ps.get("google_maps_discovered", 0),
                "google_maps_duplicates": ps.get("google_maps_duplicates", 0),
                "companyenrich_calls":    ps.get("companyenrich_calls", 0),
                "serper_calls":           ps.get("serper_calls", 0),
                "firecrawl_calls":        ps.get("firecrawl_calls", 0),
                "elapsed_seconds":        round(time.monotonic() - t_start, 1),
                "run_id":                 run_id,
            },
        )

    # ── Stage 4: Route-level dedup (within-batch) + cap ───────────────────────
    seen_websites: set[str] = set()
    unique: list[dict] = []
    for c in pipeline_companies:
        key = c.get("website", "").lower().strip().rstrip("/")
        if key and key in seen_websites:
            continue
        if key:
            seen_websites.add(key)
        else:
            key = c.get("company_name", "").lower().strip()
            if key and key in seen_websites:
                continue
            if key:
                seen_websites.add(key)
        unique.append(c)

    _log("DEDUP", f"After route-level dedup: {len(unique)} unique companies")
    await _append_log(db_early, run_id, "FILTER", "dedup",
                      f"Route-level dedup: {len(unique)} unique companies after removing batch duplicates")

    if len(unique) > target:
        unique.sort(key=lambda c: c.get("confidence", 0.0), reverse=True)
        unique = unique[:target]
        _log("DEDUP", f"Capped to target={target}")

    # ── MongoDB setup ─────────────────────────────────────────────────────────
    db        = get_db()
    coll_name = collection_for_category(industry)   # e.g. "leads_construction"
    coll      = db[coll_name]
    _log("MONGODB", f"Target collection: '{coll_name}' ({len(unique)} companies to process)")
    await _append_log(db_early, run_id, "DATABASE", "mongodb",
                      f"Saving to collection '{coll_name}' — {len(unique)} companies to process")

    # Ensure indexes exist for this category collection (non-fatal)
    try:
        await ensure_lead_indexes(db, industry)
    except Exception as _idx_exc:
        _log("MONGODB", f"Index warning (non-fatal): {_idx_exc}")

    # Register this category in the categories collection
    try:
        cats_coll = db[CATEGORIES_COLLECTION]
        await cats_coll.update_one(
            {"name": industry},
            {"$set": {"name": industry, "collection": coll_name}},
            upsert=True,
        )
    except Exception as _cat_exc:
        _log("MONGODB", f"Category register warning (non-fatal): {_cat_exc}")

    # ── Stage 5: MongoDB pre-dedup ────────────────────────────────────────────
    # Collect the identifiers we're about to process, then ask MongoDB which
    # already exist.  This is a single bulk query — not one per company.
    candidate_websites = [
        c.get("website", "").lower().strip().rstrip("/")
        for c in unique
        if c.get("website", "").strip()
    ]
    candidate_names = [
        c.get("company_name", "")
        for c in unique
        if c.get("company_name", "").strip()
    ]

    existing_websites: set[str] = set()
    existing_names:    set[str] = set()

    if candidate_websites:
        async for doc in coll.find(
            {"website": {"$in": candidate_websites}},
            {"website": 1, "_id": 0},
        ):
            w = (doc.get("website") or "").lower().strip().rstrip("/")
            if w:
                existing_websites.add(w)

    if candidate_names:
        async for doc in coll.find(
            {"company_name": {"$in": candidate_names}},
            {"company_name": 1, "_id": 0},
        ):
            n = (doc.get("company_name") or "").lower().strip()
            if n:
                existing_names.add(n)

    _log("DEDUP", (
        f"MongoDB pre-check → {len(existing_websites)} existing websites, "
        f"{len(existing_names)} existing names in '{coll_name}'"
    ))
    await _append_log(db_early, run_id, "FILTER", "db_dedup",
                      f"DB dedup check: {len(existing_websites)} existing websites, {len(existing_names)} existing names found")

    # Classify each company as NEW or DUPLICATE
    new_companies:  list[dict] = []
    dupe_companies: list[dict] = []
    for c in unique:
        w = (c.get("website") or "").lower().strip().rstrip("/")
        n = (c.get("company_name") or "").lower().strip()
        is_dup = (w and w in existing_websites) or (not w and n and n in existing_names)
        if is_dup:
            dupe_companies.append(c)
        else:
            new_companies.append(c)

    _log("DEDUP", (
        f"New leads to insert: {len(new_companies)} | "
        f"Already in DB (will update silently, NOT returned to UI): {len(dupe_companies)}"
    ))
    await _append_log(db_early, run_id, "FILTER", "classify",
                      f"Classification complete: {len(new_companies)} new leads to insert, {len(dupe_companies)} already exist in DB")

    # ── Stage 6: MongoDB upsert (ALL companies — keeps data fresh) ───────────
    now            = datetime.now(timezone.utc)
    inserted_count = 0
    updated_count  = 0

    # Track keys for newly inserted docs so we can fetch them back
    new_website_keys: list[str] = []
    new_name_keys:    list[str] = []
    inserted_doc_ids: list[str] = []  # ObjectId strings for history run

    # Social-media domain filter (applied to research_sources list)
    _SOCIAL_DOMS = frozenset({
        "instagram.com", "facebook.com", "linkedin.com", "twitter.com",
        "x.com", "youtube.com", "tiktok.com", "pinterest.com",
    })

    def _is_official_source(u: str) -> bool:
        try:
            d = urlparse(u).netloc.lower().lstrip("www.")
            return not any(d == s or d.endswith("." + s) for s in _SOCIAL_DOMS)
        except Exception:
            return True

    for company in unique:
        website        = (company.get("website") or "").strip()
        email          = company.get("email")
        company_number = company.get("company_number")
        founder_name   = company.get("founder_name")
        founder_number = company.get("founder_number")
        source_url     = company.get("source_url") or website or None
        confidence     = company.get("confidence", 0.0)
        research_src   = company.get("research_source", "serper_firecrawl")
        rsources       = list(company.get("research_sources") or [])
        rsources       = [u for u in rsources if _is_official_source(u)]
        if source_url and source_url not in rsources and _is_official_source(source_url):
            rsources = [source_url] + rsources

        has_contact   = bool(email or company_number)
        last_verified = now.isoformat() if has_contact else None
        field_verification = company.get("_field_verification") or {}

        set_fields: dict = {
            "company_name":        company.get("company_name", ""),
            "category":            industry,       # ← always store the category
            "website":             website,
            "emails":              company.get("emails", []),
            "phones":              company.get("phones", []),
            "address":             company.get("address", ""),
            "city":                company.get("city", ""),
            "state":               company.get("state", ""),
            "country":             company.get("country", ""),
            "postal_code":         company.get("postal_code", ""),
            "sources":             company.get("sources", []),
            "updated_at":          now,
            "email":               email,
            "company_number":      company_number,
            "founder_name":        founder_name,
            "founder_number":      founder_number,
            "source_url":          source_url,
            "confidence":          confidence,
            "research_source":     research_src,
            "research_sources":    rsources,
            "_field_verification": field_verification,
            # Google Maps geo fields
            "place_id":            company.get("place_id"),
            "google_maps_uri":     company.get("google_maps_uri"),
            "primary_type":        company.get("primary_type"),
            "latitude":            company.get("latitude"),
            "longitude":           company.get("longitude"),
            # People enrichment contacts (PDL → Prospeo → ContactOut merged + Origami)
            "contacts":            company.get("contacts", []),
            # Origami people array — all contacts found by Origami with tier labels
            # (structured for CRM display: Founder | CEO | Director | …)
            "people":              company.get("people", []),
            # Origami enrichment metadata
            "origami_enriched":    company.get("origami_enriched", False),
            "origami_confidence":  company.get("origami_confidence", 0.0),
            "origami_source":      company.get("origami_source", ""),
            # Founder discovery status — "found" | "found_decision_maker" |
            # "not_found" | "skipped" | "error"  (never fabricated)
            "founder_status":      company.get("founder_status", "skipped"),
            # Origami-specific founder fields (only set when sourced, never guessed)
            "founder_title":       company.get("founder_title"),
            "founder_email":       company.get("founder_email"),
            "founder_profile_url": company.get("founder_profile_url"),
        }
        if last_verified:
            set_fields["last_verified"] = last_verified

        filter_key = (
            {"website": website}
            if website
            else {"company_name": company.get("company_name", "")}
        )
        result = await coll.update_one(
            filter_key,
            {
                "$set": set_fields,
                "$setOnInsert": {
                    "created_at":       now,
                    "generation_run_id": run_id,
                    "status":           "new",   # default status for every new lead
                    "status_updated_at": now.isoformat(),
                },
            },
            upsert=True,
        )

        is_insert = bool(result.upserted_id)
        if is_insert:
            inserted_count += 1
            inserted_doc_ids.append(str(result.upserted_id))
        else:
            updated_count += 1

        # Track new insertions so we can fetch them back for the response
        if is_insert:
            if website:
                new_website_keys.append(website)
            else:
                new_name_keys.append(company.get("company_name", ""))

        _log("MONGODB", (
            f"{'INSERT' if is_insert else 'UPDATE'}: "
            f"{company.get('company_name', '?')} | "
            f"category={industry!r} | "
            f"email={'YES' if email else 'NO'} | "
            f"phone={'YES' if company_number else 'NO'} | "
            f"confidence={confidence}"
        ))

    _log("MONGODB", (
        f"Done — inserted={inserted_count} updated={updated_count} "
        f"collection='{coll_name}'"
    ))
    await _append_log(db_early, run_id, "DATABASE", "mongodb",
                      f"MongoDB upsert complete — inserted={inserted_count} updated={updated_count}")

    # Update lead_count in categories collection
    try:
        total_in_coll = await coll.count_documents({})
        await db[CATEGORIES_COLLECTION].update_one(
            {"name": industry},
            {"$set": {"lead_count": total_in_coll}},
        )
    except Exception:
        pass

    # ── Stage 7: Fetch newly inserted docs from MongoDB for the response ───────
    # IMPORTANT: we only return docs that were INSERTED this run (upserted_id set).
    # Existing/updated docs are silently refreshed but NOT shown to the user.
    if not new_website_keys and not new_name_keys:
        db_docs: list[dict] = []
        _log("RESPONSE", "No new leads inserted — returning empty list to UI")
    else:
        conditions = []
        if new_website_keys:
            conditions.append({"website": {"$in": new_website_keys}})
        if new_name_keys:
            conditions.append({"company_name": {"$in": new_name_keys}})

        fetch_filter: dict = {"$or": conditions} if len(conditions) > 1 else conditions[0]
        cursor  = coll.find(fetch_filter).sort("created_at", -1)
        db_docs = await cursor.to_list(length=inserted_count + 10)

    leads_out: list[dict] = []
    fetched_lead_ids: list[str] = []
    for doc in db_docs:
        raw_id = str(doc.get("_id", ""))
        doc["id"] = str(doc.pop("_id"))
        fetched_lead_ids.append(raw_id or doc["id"])
        for ts_field in ("created_at", "updated_at"):
            if isinstance(doc.get(ts_field), datetime):
                doc[ts_field] = doc[ts_field].isoformat()
        leads_out.append(doc)

    elapsed = round(time.monotonic() - t_start, 1)

    # ── Summary log ───────────────────────────────────────────────────────────
    n_email    = sum(1 for c in unique if c.get("email"))
    n_phone    = sum(1 for c in unique if c.get("company_number"))
    n_address  = sum(1 for c in unique if c.get("address"))
    n_founder  = sum(1 for c in unique if c.get("founder_name"))
    n_contacts = sum(1 for c in unique if c.get("contacts"))
    n_contact_emails = sum(
        sum(1 for ct in c.get("contacts", []) if ct.get("email"))
        for c in unique
    )
    n_contact_phones = sum(
        sum(1 for ct in c.get("contacts", []) if ct.get("phone"))
        for c in unique
    )
    _log("LEADS", (
        f"COMPLETE in {elapsed}s | "
        f"pipeline_total={len(unique)} | "
        f"new_inserted={inserted_count} | "
        f"already_existed={updated_count} | "
        f"returned_to_ui={len(leads_out)} | "
        f"email={n_email}/{len(unique)} | "
        f"phone={n_phone}/{len(unique)} | "
        f"address={n_address}/{len(unique)} | "
        f"founder={n_founder}/{len(unique)} | "
        f"contacts={n_contacts}/{len(unique)} | "
        f"contact_emails={n_contact_emails} | "
        f"contact_phones={n_contact_phones} | "
        f"origami_calls={ps.get('origami_calls', 0)} | "
        f"origami_contacts={ps.get('origami_contacts_found', 0)} | "
        f"origami_founders={ps.get('origami_founders_found', 0)} | "
        f"origami_emails={ps.get('origami_emails_found', 0)} | "
        f"gmaps_discovered={ps.get('google_maps_discovered', 0)} | "
        f"gmaps_dupes={ps.get('google_maps_duplicates', 0)} | "
        f"ce_calls={ps.get('companyenrich_calls', 0)} | "
        f"serper_calls={ps.get('serper_calls', 0)} | "
        f"firecrawl_calls={ps.get('firecrawl_calls', 0)}"
    ))

    # ── API Contribution Report ───────────────────────────────────────────────
    # Prints a detailed per-API breakdown to backend logs after every run.
    # READ-ONLY — never modifies pipeline data or lead dicts.
    # API keys are NEVER read or logged here.
    try:
        from app.services.api_contribution_logger import print_contribution_report
        print_contribution_report(unique, ps)
    except Exception as _cr_exc:
        _log("LEADS", f"API contribution report error (non-fatal): {_cr_exc}")

    final_pipeline_stats = {
        "google_maps_discovered":      ps.get("google_maps_discovered", 0),
        "google_maps_duplicates":      ps.get("google_maps_duplicates", 0),
        "companyenrich_calls":         ps.get("companyenrich_calls", 0),
        "companyenrich_fields_filled": ps.get("companyenrich_fields_filled", 0),
        "serper_calls":                ps.get("serper_calls", 0),
        "serper_fields_filled":        ps.get("serper_fields_filled", 0),
        "firecrawl_calls":             ps.get("firecrawl_calls", 0),
        "firecrawl_fields_filled":     ps.get("firecrawl_fields_filled", 0),
        "final_valid_companies":       ps.get("final_valid_companies", len(unique)),
        "db_dedup_skipped":            len(dupe_companies),
        # Origami enrichment stats
        "origami_calls":               ps.get("origami_calls", 0),
        "origami_contacts_found":      ps.get("origami_contacts_found", 0),
        "origami_founders_found":      ps.get("origami_founders_found", 0),
        "origami_emails_found":        ps.get("origami_emails_found", 0),
        "origami_skipped":             ps.get("origami_skipped", 0),
        # People enrichment orchestrator
        "people_companies_processed":  ps.get("people_companies_processed", 0),
        "people_contacts_found":       ps.get("people_contacts_found", 0),
        "people_emails_found":         ps.get("people_emails_found", 0),
        "people_phones_found":         ps.get("people_phones_found", 0),
        "people_target_reached":       ps.get("people_target_reached", 0),
        "people_auth_failures":        ps.get("people_auth_failures", 0),
        "pdl_calls":                   ps.get("pdl_calls", 0),
        "pdl_contacts":                ps.get("pdl_contacts", 0),
        "prospeo_calls":               ps.get("prospeo_calls", 0),
        "prospeo_contacts":            ps.get("prospeo_contacts", 0),
        "contactout_calls":            ps.get("contactout_calls", 0),
        "contactout_contacts":         ps.get("contactout_contacts", 0),
        # Legacy compat
        "pdl_companies_searched":      ps.get("pdl_companies_searched", 0),
        "pdl_contacts_found":          ps.get("pdl_contacts_found", 0),
        "pdl_emails_found":            ps.get("pdl_emails_found", 0),
        "pdl_phones_found":            ps.get("pdl_phones_found", 0),
        "pdl_api_calls":               ps.get("pdl_api_calls", 0),
        "pdl_auth_failures":           ps.get("pdl_auth_failures", 0),
        "elapsed_seconds":             elapsed,
        "run_id":                      run_id,
    }

    # ── Complete the history run ───────────────────────────────────────────────
    run_statistics = {
        "companies_discovered": ps.get("google_maps_discovered", len(pipeline_companies)),
        "companies_processed":  len(unique),
        "leads_generated":      inserted_count,
        "duplicates":           len(dupe_companies),
        "rejected":             max(0, len(pipeline_companies) - len(unique)),
        "errors":               0,
        "with_email":           n_email,
        "with_phone":           n_phone,
        "with_founder":         n_founder,
        "contacts_found":       n_contacts,
        "companyenrich_calls":  ps.get("companyenrich_calls", 0),
        "serper_calls":         ps.get("serper_calls", 0),
        "firecrawl_calls":      ps.get("firecrawl_calls", 0),
        "elapsed_seconds":      elapsed,
    }
    completion_logs = [
        {"level": "COMPLETE", "stage": "pipeline",
         "message": f"{inserted_count} new leads generated, {updated_count} existing updated"},
        {"level": "COMPLETE", "stage": "pipeline",
         "message": f"Generation completed in {elapsed}s"},
    ]
    await _complete_history_run(
        db_early, run_id,
        inserted=inserted_count,
        updated=updated_count,
        lead_ids=inserted_doc_ids or fetched_lead_ids,
        pipeline_stats=final_pipeline_stats,
        stats=run_statistics,
        t_start=t_start,
        logs_extra=completion_logs,
    )
    _log("HISTORY", f"Run {run_id} completed — inserted={inserted_count}")

    return MongoLeadsResponse(
        success=True,
        inserted=inserted_count,
        updated=updated_count,
        total=len(leads_out),
        query=query,
        timestamp=datetime.now(timezone.utc).isoformat(),
        leads=leads_out,
        pipeline_stats=final_pipeline_stats,
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /leads/today
# Must be declared BEFORE GET /leads/{lead_id} to avoid path-param conflict.
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/leads/today",
    summary="Today's generated leads across ALL category collections",
    response_model=dict,
    tags=["Leads"],
)
async def get_today_leads(
    category: Optional[str] = Query(None, description="Limit to one category collection (omit = all categories)"),
    per_page: int            = Query(100, ge=1, le=500, description="Max leads to return per category"),
):
    """
    Return all leads whose created_at falls within today (UTC midnight → 23:59:59).

    When no category is supplied this endpoint fans out across EVERY known
    category collection (leads_construction, leads_real_estate, …) plus the
    legacy 'leads' collection, merges the results, and returns them sorted
    newest-first.

    Response shape:
        {
          "success": true,
          "date":    "YYYY-MM-DD",
          "total":   N,
          "by_category": [{"category": "...", "count": N}, ...],
          "leads":   [...],          // up to per_page * num_categories, newest first
          "summary": {
            "with_email": N, "with_phone": N, "with_founder": N,
            "reddit": N, "maps": N,
          }
        }
    """
    import asyncio as _asyncio
    from datetime import date as _date

    db        = get_db()
    today_str = _date.today().isoformat()          # "YYYY-MM-DD"
    date_filter = {
        "$gte": f"{today_str}T00:00:00",
        "$lte": f"{today_str}T23:59:59",
    }
    mongo_filter: dict = {"created_at": date_filter}

    # ── Single-category fast path ─────────────────────────────────────────────
    if category:
        coll_name = collection_for_category(category)
        coll      = db[coll_name]
        total     = await coll.count_documents(mongo_filter)
        cursor    = coll.find(mongo_filter).sort("created_at", -1).limit(per_page)
        raw_docs  = await cursor.to_list(length=per_page)
        leads_out = []
        for doc in raw_docs:
            doc["id"] = str(doc.pop("_id"))
            for tf in ("created_at", "updated_at"):
                if isinstance(doc.get(tf), datetime):
                    doc[tf] = doc[tf].isoformat()
            leads_out.append(doc)
        summary = {
            "with_email":   sum(1 for l in leads_out if l.get("email")),
            "with_phone":   sum(1 for l in leads_out if l.get("company_number") or l.get("phones")),
            "with_founder": sum(1 for l in leads_out if l.get("founder_name")),
            "reddit":       sum(1 for l in leads_out if l.get("research_source") == "reddit"),
            "maps":         sum(1 for l in leads_out if l.get("research_source") != "reddit"),
        }
        return {
            "success":     True,
            "date":        today_str,
            "total":       total,
            "by_category": [{"category": category, "count": total}],
            "leads":       leads_out,
            "summary":     summary,
        }

    # ── All-categories fan-out ────────────────────────────────────────────────
    # Gather known category names from the categories collection
    try:
        cats_coll  = db[CATEGORIES_COLLECTION]
        cat_cursor = cats_coll.find({}, {"name": 1, "_id": 0})
        cat_docs   = await cat_cursor.to_list(length=500)
        cat_names  = [d["name"] for d in cat_docs if d.get("name")]
    except Exception:
        cat_names = []

    if not cat_names:
        cat_names = ALL_CATEGORIES

    # Build collection list: all category slugs + legacy fallback
    coll_map: dict[str, str] = {}                   # coll_name → display category
    for cat in cat_names:
        coll_map[collection_for_category(cat)] = cat
    coll_map[COLLECTION_NAME] = "Legacy"             # fallback 'leads' collection

    # Fan out across all collections concurrently
    async def _query_one(coll_name: str, display_cat: str):
        try:
            coll  = db[coll_name]
            total = await coll.count_documents(mongo_filter)
            if total == 0:
                return display_cat, 0, []
            cursor = coll.find(mongo_filter).sort("created_at", -1).limit(per_page)
            docs   = await cursor.to_list(length=per_page)
            out    = []
            for doc in docs:
                doc["id"] = str(doc.pop("_id"))
                for tf in ("created_at", "updated_at"):
                    if isinstance(doc.get(tf), datetime):
                        doc[tf] = doc[tf].isoformat()
                if not doc.get("category"):
                    doc["category"] = display_cat
                out.append(doc)
            return display_cat, total, out
        except Exception:
            return display_cat, 0, []

    gathered = await _asyncio.gather(
        *[_query_one(cn, cat) for cn, cat in coll_map.items()]
    )

    # Merge + sort all docs newest-first
    all_leads: list[dict] = []
    by_category: list[dict] = []
    grand_total = 0

    for cat_label, count, docs in gathered:
        if count > 0:
            by_category.append({"category": cat_label, "count": count})
            grand_total += count
            all_leads.extend(docs)

    # Sort merged list by created_at descending (ISO strings sort correctly)
    all_leads.sort(key=lambda d: d.get("created_at") or "", reverse=True)
    all_leads = all_leads[:per_page]   # cap final result

    summary = {
        "with_email":   sum(1 for l in all_leads if l.get("email")),
        "with_phone":   sum(1 for l in all_leads if l.get("company_number") or l.get("phones")),
        "with_founder": sum(1 for l in all_leads if l.get("founder_name")),
        "reddit":       sum(1 for l in all_leads if l.get("research_source") == "reddit"),
        "maps":         sum(1 for l in all_leads if l.get("research_source") != "reddit"),
    }

    return {
        "success":     True,
        "date":        today_str,
        "total":       grand_total,
        "by_category": sorted(by_category, key=lambda x: x["count"], reverse=True),
        "leads":       all_leads,
        "summary":     summary,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /leads
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/leads",
    summary="List stored leads from MongoDB",
    response_model=MongoLeadsResponse,
    tags=["Leads"],
)
async def get_leads(
    category:  Optional[str] = Query(None, description="Filter by industry category"),
    search:    Optional[str] = Query(None, description="Search across company_name, email, phones, founder_name"),
    status:    Optional[str] = Query(None, description="Filter by status: new|interested|not_interested"),
    tab:       Optional[str] = Query(None, description="Smart tab: new_leads|old_untouched|interested|not_interested|follow_ups|all"),
    date_from: Optional[str] = Query(None, description="Filter created_at >= YYYY-MM-DD"),
    date_to:   Optional[str] = Query(None, description="Filter created_at <= YYYY-MM-DD"),
    page:      int           = Query(1,   ge=1),
    per_page:  int           = Query(100, ge=1, le=500),
):
    """
    Fetch stored leads from MongoDB with server-side filtering.
    Delegates filter building to _build_leads_filter().
    """
    db        = get_db()
    coll_name = collection_for_category(category) if category else COLLECTION_NAME
    coll      = db[coll_name]

    mongo_filter = _build_leads_filter(tab, status, search, date_from, date_to)

    total   = await coll.count_documents(mongo_filter)
    skip    = (page - 1) * per_page
    cursor  = coll.find(mongo_filter).sort("created_at", -1).skip(skip).limit(per_page)
    db_docs = await cursor.to_list(length=per_page)

    leads_out = []
    for doc in db_docs:
        doc["id"] = str(doc.pop("_id"))
        for ts_field in ("created_at", "updated_at"):
            if isinstance(doc.get(ts_field), datetime):
                doc[ts_field] = doc[ts_field].isoformat()
        leads_out.append(doc)

    return MongoLeadsResponse(
        success=True,
        inserted=0,
        updated=0,
        total=total,
        query=category or "",
        timestamp=datetime.now(timezone.utc).isoformat(),
        leads=leads_out,
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /leads/status-counts  — MUST be before /leads/{lead_id} so FastAPI does
# not treat the literal string "status-counts" as a lead_id wildcard.
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/leads/status-counts",
    summary="Count leads by status + smart tabs for a category",
    response_model=dict,
    tags=["Leads"],
)
async def get_status_counts(
    category: Optional[str] = Query(None, description="Industry category — routes to correct collection"),
):
    """
    Returns MongoDB counts for all tab views:
      { new: N, old_untouched: N, interested: N, not_interested: N,
        follow_ups: N, total: N }

    new_leads     = status in [null/"new"] AND created_at >= now-2d
    old_untouched = status in [null/"new"] AND created_at <  now-2d
    interested    = status == "interested"
    not_interested= status == "not_interested"
    follow_ups    = follow_up_date is set (not null/empty)
    total         = all documents
    """
    from datetime import timedelta

    db        = get_db()
    coll_name = collection_for_category(category) if category else COLLECTION_NAME
    coll      = db[coll_name]

    cutoff_2d = (datetime.now(timezone.utc) - timedelta(days=2)).isoformat()

    new_status_filter = {"$or": [
        {"status": "new"},
        {"status": {"$exists": False}},
        {"status": None},
    ]}

    # Run all counts in parallel-ish (sequential awaits, all fast index hits)
    total            = await coll.count_documents({})
    interested       = await coll.count_documents({"status": "interested"})
    not_interested   = await coll.count_documents({"status": "not_interested"})
    new_leads        = await coll.count_documents({"$and": [new_status_filter, {"created_at": {"$gte": cutoff_2d}}]})
    old_untouched    = await coll.count_documents({"$and": [new_status_filter, {"created_at": {"$lt":  cutoff_2d}}]})
    follow_ups       = await coll.count_documents({"follow_up_date": {"$nin": [None, ""]}})

    return {
        "success": True,
        "category": category,
        "counts": {
            "new":           new_leads,
            "old_untouched": old_untouched,
            "interested":    interested,
            "not_interested": not_interested,
            "follow_ups":    follow_ups,
            "total":         total,
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /leads/follow-ups  — MUST be before /leads/{lead_id} so FastAPI does
# not treat the literal string "follow-ups" as a lead_id wildcard.
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/leads/follow-ups",
    summary="Get today's, overdue, and upcoming follow-up leads across all collections",
    response_model=dict,
    tags=["Leads"],
)
async def get_follow_ups(
    category: Optional[str] = Query(None, description="Limit to one category collection"),
):
    """
    Returns three lists:
      - overdue:  leads with follow_up_date < today AND status != 'not_interested'
                  AND follow_up_completed != True
      - today:    leads with follow_up_date == today AND follow_up_completed != True
      - upcoming: leads with follow_up_date > today (next 30 days)
                  AND follow_up_completed != True

    Excludes leads marked as not_interested.
    Excludes leads where follow_up_completed == True.
    """
    from datetime import date, timedelta

    db       = get_db()
    today    = date.today().isoformat()          # "YYYY-MM-DD"
    in_30    = (date.today() + timedelta(days=30)).isoformat()

    # Base filter: exclude completed follow-ups
    not_completed = {"$or": [
        {"follow_up_completed": {"$exists": False}},
        {"follow_up_completed": False},
        {"follow_up_completed": None},
    ]}

    # Decide which collections to scan
    if category:
        colls = [collection_for_category(category)]
    else:
        cats_cursor = db[CATEGORIES_COLLECTION].find({}, {"collection": 1, "_id": 0})
        cats_docs   = await cats_cursor.to_list(length=500)
        colls = list({d["collection"] for d in cats_docs if d.get("collection")})
        if not colls:
            colls = [COLLECTION_NAME]

    # Projection — include founder_name and company_number for richer notifications
    projection = {
        "company_name": 1, "founder_name": 1,
        "email": 1, "company_number": 1, "founder_number": 1,
        "follow_up_date": 1, "follow_up_completed": 1,
        "status": 1, "category": 1, "_id": 1,
    }

    overdue_leads: list[dict] = []
    today_leads:   list[dict] = []
    upcoming_leads: list[dict] = []

    for cname in colls:
        coll = db[cname]

        # ── Overdue: date < today, not not_interested, not completed ─────────
        overdue_cursor = coll.find(
            {"$and": [
                {"follow_up_date": {"$nin": [None, ""], "$lt": today}},
                {"status": {"$ne": "not_interested"}},
                not_completed,
            ]},
            projection,
        ).sort("follow_up_date", 1)
        async for doc in overdue_cursor:
            doc["id"] = str(doc.pop("_id"))
            overdue_leads.append(doc)

        # ── Today + upcoming: date >= today <= today+30, not completed ────────
        future_cursor = coll.find(
            {"$and": [
                {"follow_up_date": {"$gte": today, "$lte": in_30}},
                {"status": {"$ne": "not_interested"}},
                not_completed,
            ]},
            projection,
        ).sort("follow_up_date", 1)
        async for doc in future_cursor:
            doc["id"] = str(doc.pop("_id"))
            if doc.get("follow_up_date") == today:
                today_leads.append(doc)
            else:
                upcoming_leads.append(doc)

    # Sort overdue with most overdue first
    overdue_leads.sort(key=lambda d: d.get("follow_up_date") or "")

    return {
        "success":        True,
        "overdue":        overdue_leads,
        "today":          today_leads,
        "upcoming":       upcoming_leads,
        "overdue_count":  len(overdue_leads),
        "today_count":    len(today_leads),
        "upcoming_count": len(upcoming_leads),
        # total actionable = overdue + today
        "due_count":      len(overdue_leads) + len(today_leads),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Manual CRUD
# ─────────────────────────────────────────────────────────────────────────────

@router.post(
    "/leads",
    summary="Create a new lead manually",
    response_model=LeadResponse,
    status_code=status.HTTP_201_CREATED,
    tags=["Leads"],
)
def create_lead(payload: LeadCreateRequest):
    return LeadController.create_lead(payload)


@router.get(
    "/leads/{lead_id}",
    summary="Get a single lead",
    response_model=LeadResponse,
    tags=["Leads"],
)
def get_lead(lead_id: str):
    try:
        return LeadController.get_lead(lead_id)
    except LeadNotFoundError as exc:
        raise _handle_not_found(exc)


@router.patch(
    "/leads/{lead_id}",
    summary="Partially update a lead",
    response_model=LeadResponse,
    tags=["Leads"],
)
def update_lead(lead_id: str, payload: LeadUpdateRequest):
    try:
        return LeadController.update_lead(lead_id, payload)
    except LeadNotFoundError as exc:
        raise _handle_not_found(exc)


@router.delete(
    "/leads/{lead_id}",
    summary="Delete a lead",
    response_model=MessageResponse,
    tags=["Leads"],
)
def delete_lead(lead_id: str):
    try:
        return LeadController.delete_lead(lead_id)
    except LeadNotFoundError as exc:
        raise _handle_not_found(exc)


# ─────────────────────────────────────────────────────────────────────────────
# PATCH /leads/{lead_id}/status
# ─────────────────────────────────────────────────────────────────────────────

from pydantic import BaseModel as _StatusBodyBase


class _StatusUpdateBody(_StatusBodyBase):
    status:   str
    category: Optional[str] = None


class _NoteBody(_StatusBodyBase):
    text:     str
    category: Optional[str] = None


class _FollowUpBody(_StatusBodyBase):
    follow_up_date: Optional[str] = None   # ISO date string "YYYY-MM-DD" or null to clear
    category:       Optional[str] = None


class _FollowUpCompleteBody(_StatusBodyBase):
    category: Optional[str] = None


# The three allowed statuses for the lead status management feature.
# Leads are generated with status = "new" by default.
LEAD_STATUSES = {"new", "interested", "not_interested"}


@router.patch(
    "/leads/{lead_id}/status",
    summary="Update CRM status for a lead (new | interested | not_interested)",
    response_model=dict,
    tags=["Leads"],
)
async def update_lead_status(lead_id: str, payload: _StatusUpdateBody):
    """
    Update the status field of an existing lead document.

    Allowed values: new | interested | not_interested

    - Finds the lead by MongoDB ObjectId across the correct category collection.
    - Writes  status = payload.status  and  status_updated_at = utcnow.
    - Returns the full updated lead document so the UI can update immediately.
    - Never creates a duplicate lead.
    """
    from bson import ObjectId
    from bson.errors import InvalidId

    if payload.status not in LEAD_STATUSES:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid status '{payload.status}'. Allowed: {sorted(LEAD_STATUSES)}",
        )

    db        = get_db()
    now       = datetime.now(timezone.utc)

    # Route to the correct per-category collection (or legacy fallback)
    coll_name = collection_for_category(payload.category) if payload.category else None

    # Try to build an ObjectId from lead_id
    try:
        oid = ObjectId(lead_id)
    except (InvalidId, Exception):
        oid = None

    # Search strategy:
    #  1. If category given → look only in that collection
    #  2. Otherwise → try the legacy "leads" collection first, then scan all known categories
    async def _try_update(cname: str):
        coll   = db[cname]
        flt    = {"_id": oid} if oid else {"id": lead_id}
        result = await coll.update_one(
            flt,
            {"$set": {
                "status":            payload.status,
                "status_updated_at": now.isoformat(),
                "updated_at":        now,
            }},
        )
        if result.matched_count > 0:
            doc = await coll.find_one(flt)
            return doc
        return None

    updated_doc = None

    if coll_name:
        updated_doc = await _try_update(coll_name)
    else:
        # Try legacy collection first
        updated_doc = await _try_update(COLLECTION_NAME)
        if not updated_doc:
            # Walk all known category collections
            cats_cursor = db[CATEGORIES_COLLECTION].find(
                {}, {"collection": 1, "_id": 0}
            )
            async for cat in cats_cursor:
                c = cat.get("collection")
                if c and c != COLLECTION_NAME:
                    updated_doc = await _try_update(c)
                    if updated_doc:
                        break

    if not updated_doc:
        raise HTTPException(
            status_code=404,
            detail=f"Lead '{lead_id}' not found. Pass 'category' to route to the correct collection.",
        )

    # Serialize ObjectId → string
    updated_doc["id"] = str(updated_doc.pop("_id"))
    for ts in ("created_at", "updated_at"):
        if isinstance(updated_doc.get(ts), datetime):
            updated_doc[ts] = updated_doc[ts].isoformat()

    return {
        "success":          True,
        "lead_id":          lead_id,
        "status":           payload.status,
        "status_updated_at": now.isoformat(),
        "lead":             updated_doc,
    }


# ─────────────────────────────────────────────────────────────────────────────
# POST /leads/{lead_id}/notes
# ─────────────────────────────────────────────────────────────────────────────

@router.post(
    "/leads/{lead_id}/notes",
    summary="Add a note to an existing lead",
    response_model=dict,
    tags=["Leads"],
)
async def add_lead_note(lead_id: str, payload: _NoteBody):
    """
    Append a note to the lead's `notes` array inside the existing document.
    Never creates a new lead document.
    """
    from bson import ObjectId
    from bson.errors import InvalidId

    text = (payload.text or "").strip()
    if not text:
        raise HTTPException(status_code=422, detail="Note text cannot be empty.")
    if len(text) > 2000:
        raise HTTPException(status_code=422, detail="Note text must be ≤ 2000 characters.")

    db  = get_db()
    now = datetime.now(timezone.utc)
    note = {"text": text, "created_at": now.isoformat()}

    try:
        oid = ObjectId(lead_id)
    except (InvalidId, Exception):
        oid = None

    coll_name = collection_for_category(payload.category) if payload.category else None

    async def _try_add(cname: str):
        coll = db[cname]
        flt  = {"_id": oid} if oid else {"id": lead_id}
        result = await coll.update_one(
            flt,
            {"$push": {"notes": note}, "$set": {"updated_at": now}},
        )
        if result.matched_count > 0:
            return await coll.find_one(flt)
        return None

    doc = None
    if coll_name:
        doc = await _try_add(coll_name)
    else:
        doc = await _try_add(COLLECTION_NAME)
        if not doc:
            async for cat in db[CATEGORIES_COLLECTION].find({}, {"collection": 1, "_id": 0}):
                c = cat.get("collection")
                if c and c != COLLECTION_NAME:
                    doc = await _try_add(c)
                    if doc:
                        break

    if not doc:
        raise HTTPException(status_code=404, detail=f"Lead '{lead_id}' not found.")

    doc["id"] = str(doc.pop("_id"))
    for ts in ("created_at", "updated_at"):
        if isinstance(doc.get(ts), datetime):
            doc[ts] = doc[ts].isoformat()

    return {"success": True, "note": note, "lead": doc}


# ─────────────────────────────────────────────────────────────────────────────
# PATCH /leads/{lead_id}/follow-up
# ─────────────────────────────────────────────────────────────────────────────

@router.patch(
    "/leads/{lead_id}/follow-up",
    summary="Set or clear the follow-up date for a lead",
    response_model=dict,
    tags=["Leads"],
)
async def update_follow_up(lead_id: str, payload: _FollowUpBody):
    """
    Set `follow_up_date` (ISO date string "YYYY-MM-DD") or clear it (null).
    Validates the date format when provided.
    """
    from bson import ObjectId
    from bson.errors import InvalidId
    import re as _re

    fud = (payload.follow_up_date or "").strip() or None
    if fud and not _re.match(r'^\d{4}-\d{2}-\d{2}$', fud):
        raise HTTPException(status_code=422, detail="follow_up_date must be YYYY-MM-DD or null.")

    db  = get_db()
    now = datetime.now(timezone.utc)

    try:
        oid = ObjectId(lead_id)
    except (InvalidId, Exception):
        oid = None

    coll_name = collection_for_category(payload.category) if payload.category else None

    async def _try_update(cname: str):
        coll = db[cname]
        flt  = {"_id": oid} if oid else {"id": lead_id}
        result = await coll.update_one(
            flt,
            {"$set": {"follow_up_date": fud, "updated_at": now}},
        )
        if result.matched_count > 0:
            return await coll.find_one(flt)
        return None

    doc = None
    if coll_name:
        doc = await _try_update(coll_name)
    else:
        doc = await _try_update(COLLECTION_NAME)
        if not doc:
            async for cat in db[CATEGORIES_COLLECTION].find({}, {"collection": 1, "_id": 0}):
                c = cat.get("collection")
                if c and c != COLLECTION_NAME:
                    doc = await _try_update(c)
                    if doc:
                        break

    if not doc:
        raise HTTPException(status_code=404, detail=f"Lead '{lead_id}' not found.")

    doc["id"] = str(doc.pop("_id"))
    for ts in ("created_at", "updated_at"):
        if isinstance(doc.get(ts), datetime):
            doc[ts] = doc[ts].isoformat()

    return {"success": True, "follow_up_date": fud, "lead": doc}


# ─────────────────────────────────────────────────────────────────────────────
# PATCH /leads/{lead_id}/follow-up-complete
# Mark a follow-up as completed. Sets follow_up_completed=True and clears
# the follow_up_date so the lead disappears from Due Today / Overdue.
# Does NOT delete the lead.
# ─────────────────────────────────────────────────────────────────────────────

class _FollowUpCompleteBody(_StatusBodyBase):
    category: Optional[str] = None


@router.patch(
    "/leads/{lead_id}/follow-up-complete",
    summary="Mark a follow-up as completed (removes from Due Today / Overdue)",
    response_model=dict,
    tags=["Leads"],
)
async def complete_follow_up(lead_id: str, payload: _FollowUpCompleteBody):
    """
    Sets follow_up_completed = True and clears follow_up_date on the lead.
    The lead is NOT deleted — it remains in the CRM with its existing status.
    After this call the lead will no longer appear in the follow-ups lists.
    """
    from bson import ObjectId
    from bson.errors import InvalidId

    db  = get_db()
    now = datetime.now(timezone.utc)

    try:
        oid = ObjectId(lead_id)
    except (InvalidId, Exception):
        oid = None

    coll_name = collection_for_category(payload.category) if payload.category else None

    async def _try_complete(cname: str):
        coll   = db[cname]
        flt    = {"_id": oid} if oid else {"id": lead_id}
        result = await coll.update_one(
            flt,
            {"$set": {
                "follow_up_completed": True,
                "follow_up_date":      None,
                "updated_at":          now,
            }},
        )
        if result.matched_count > 0:
            return await coll.find_one(flt)
        return None

    doc = None
    if coll_name:
        doc = await _try_complete(coll_name)
    else:
        doc = await _try_complete(COLLECTION_NAME)
        if not doc:
            async for cat in db[CATEGORIES_COLLECTION].find({}, {"collection": 1, "_id": 0}):
                c = cat.get("collection")
                if c and c != COLLECTION_NAME:
                    doc = await _try_complete(c)
                    if doc:
                        break

    if not doc:
        raise HTTPException(status_code=404, detail=f"Lead '{lead_id}' not found.")

    doc["id"] = str(doc.pop("_id"))
    for ts in ("created_at", "updated_at"):
        if isinstance(doc.get(ts), datetime):
            doc[ts] = doc[ts].isoformat()

    return {"success": True, "lead_id": lead_id, "lead": doc}


# ─────────────────────────────────────────────────────────────────────────────
# Shared helper — build a MongoDB filter dict from the standard query params
# (reused by GET /leads, GET /leads/export/csv, GET /leads/export/excel)
# ─────────────────────────────────────────────────────────────────────────────

def _build_leads_filter(
    tab:       Optional[str],
    status:    Optional[str],
    search:    Optional[str],
    date_from: Optional[str],
    date_to:   Optional[str],
) -> dict:
    """Return a MongoDB filter dict from the standard query params."""
    from datetime import timedelta

    mongo_filter: dict = {}
    cutoff_2d = (datetime.now(timezone.utc) - timedelta(days=2)).isoformat()

    new_status_filter = {"$or": [
        {"status": "new"},
        {"status": {"$exists": False}},
        {"status": None},
    ]}

    if tab == "new_leads":
        mongo_filter["$and"] = [
            new_status_filter,
            {"created_at": {"$gte": cutoff_2d}},
        ]
    elif tab == "old_untouched":
        mongo_filter["$and"] = [
            new_status_filter,
            {"created_at": {"$lt": cutoff_2d}},
        ]
    elif tab == "interested":
        mongo_filter["status"] = "interested"
    elif tab == "not_interested":
        mongo_filter["status"] = "not_interested"
    elif tab == "follow_ups":
        mongo_filter["follow_up_date"] = {"$nin": [None, ""]}
    elif status:
        if status == "new":
            mongo_filter.update(new_status_filter)
        else:
            mongo_filter["status"] = status

    # Date range
    if date_from or date_to:
        date_filter: dict = {}
        if date_from:
            date_filter["$gte"] = f"{date_from}T00:00:00"
        if date_to:
            date_filter["$lte"] = f"{date_to}T23:59:59"
        existing_and = mongo_filter.get("$and")
        if existing_and:
            existing_and.append({"created_at": date_filter})
        else:
            existing_ca = mongo_filter.get("created_at")
            if existing_ca:
                mongo_filter["$and"] = [{"created_at": existing_ca}, {"created_at": date_filter}]
                del mongo_filter["created_at"]
            else:
                mongo_filter["created_at"] = date_filter

    # Multi-field search
    if search and search.strip():
        q = search.strip()
        search_cond = {"$or": [
            {"company_name":   {"$regex": q, "$options": "i"}},
            {"email":          {"$regex": q, "$options": "i"}},
            {"emails":         {"$regex": q, "$options": "i"}},
            {"company_number": {"$regex": q, "$options": "i"}},
            {"phones":         {"$regex": q, "$options": "i"}},
            {"founder_name":   {"$regex": q, "$options": "i"}},
            {"founder_number": {"$regex": q, "$options": "i"}},
            {"address":        {"$regex": q, "$options": "i"}},
        ]}
        existing_and = mongo_filter.get("$and")
        if existing_and:
            existing_and.append(search_cond)
        elif "$or" in mongo_filter:
            mongo_filter["$and"] = [{"$or": mongo_filter.pop("$or")}, search_cond]
        else:
            mongo_filter.update(search_cond)

    return mongo_filter


def _doc_to_export_row(doc: dict) -> dict:
    """Flatten a MongoDB lead document into a flat export row."""
    # Resolve email
    email = doc.get("email") or ""
    if not email and doc.get("emails"):
        email = doc["emails"][0] if isinstance(doc["emails"], list) else str(doc["emails"])

    # Resolve phone
    phone = doc.get("company_number") or ""
    if not phone and doc.get("phones"):
        phone = doc["phones"][0] if isinstance(doc["phones"], list) else str(doc["phones"])

    # Flatten notes → one cell
    notes_list = doc.get("notes") or []
    if notes_list:
        parts = []
        for n in notes_list:
            ts  = n.get("created_at", "")[:10] if isinstance(n, dict) else ""
            txt = n.get("text", "") if isinstance(n, dict) else str(n)
            parts.append(f"[{ts}] {txt}" if ts else txt)
        notes_str = " | ".join(parts)
    else:
        notes_str = ""

    # created_at — keep only date part
    raw_created = doc.get("created_at") or ""
    created_date = str(raw_created)[:10] if raw_created else ""

    return {
        "Name":          doc.get("founder_name") or "",
        "Email":         email,
        "Phone":         phone,
        "Company":       doc.get("company_name") or "",
        "Designation":   doc.get("designation") or "",
        "Category":      doc.get("category") or "",
        "Platform":      doc.get("platform") or doc.get("research_source") or "",
        "Form":          doc.get("form_name") or doc.get("generation_run_id") or "",
        "Status":        doc.get("status") or "new",
        "Notes":         notes_str,
        "Follow-up Date": doc.get("follow_up_date") or "",
        "Created Date":  created_date,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /leads/export/csv  — MUST be before /leads/{lead_id}
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/leads/export/csv",
    summary="Export filtered leads as CSV",
    tags=["Leads"],
)
async def export_leads_csv(
    category:  Optional[str] = Query(None),
    tab:       Optional[str] = Query(None),
    status:    Optional[str] = Query(None),
    search:    Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to:   Optional[str] = Query(None),
):
    """Stream a CSV file of all leads matching the current filters (no pagination limit)."""
    import csv
    import io
    from fastapi.responses import StreamingResponse

    db        = get_db()
    coll_name = collection_for_category(category) if category else COLLECTION_NAME
    coll      = db[coll_name]

    mongo_filter = _build_leads_filter(tab, status, search, date_from, date_to)
    cursor = coll.find(mongo_filter).sort("created_at", -1)
    docs   = await cursor.to_list(length=10000)

    output = io.StringIO()
    columns = ["Name","Email","Phone","Company","Designation","Category",
               "Platform","Form","Status","Notes","Follow-up Date","Created Date"]
    writer = csv.DictWriter(output, fieldnames=columns, lineterminator="\n")
    writer.writeheader()
    for doc in docs:
        doc["id"] = str(doc.pop("_id", ""))
        writer.writerow(_doc_to_export_row(doc))

    output.seek(0)
    filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /leads/export/excel  — MUST be before /leads/{lead_id}
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/leads/export/excel",
    summary="Export filtered leads as Excel (.xlsx)",
    tags=["Leads"],
)
async def export_leads_excel(
    category:  Optional[str] = Query(None),
    tab:       Optional[str] = Query(None),
    status:    Optional[str] = Query(None),
    search:    Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to:   Optional[str] = Query(None),
):
    """Stream an XLSX file of all leads matching the current filters."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    db        = get_db()
    coll_name = collection_for_category(category) if category else COLLECTION_NAME
    coll      = db[coll_name]

    mongo_filter = _build_leads_filter(tab, status, search, date_from, date_to)
    cursor = coll.find(mongo_filter).sort("created_at", -1)
    docs   = await cursor.to_list(length=10000)

    columns = ["Name","Email","Phone","Company","Designation","Category",
               "Platform","Form","Status","Notes","Follow-up Date","Created Date"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header row styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, doc in enumerate(docs, start=2):
        doc["id"] = str(doc.pop("_id", ""))
        row = _doc_to_export_row(doc)
        for col_idx, col_name in enumerate(columns, start=1):
            val  = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths (cap at 60)
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len    = max(
            len(col_name),
            *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, len(docs) + 2))
        ) if docs else len(col_name)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /leads/export/excel/all-categories
# Multi-sheet Excel: one sheet per category that has leads + one summary sheet.
# Single click downloads every lead across every category collection.
# MUST be before /leads/{lead_id}
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/leads/export/excel/all-categories",
    summary="Export ALL leads from ALL categories as a multi-sheet Excel file (one sheet per category)",
    tags=["Leads"],
)
async def export_all_categories_excel():
    """
    Iterates every known category collection in MongoDB and produces a single
    .xlsx file where:
      - Sheet 1: "All Leads"  — every lead from every category merged together
      - Sheet 2+: one sheet per category that has at least one lead

    All sheets use the same indigo header style as the single-category export.
    Column widths are auto-fitted per sheet (capped at 60).
    The file is streamed directly to the browser so it downloads in one click.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    db = get_db()

    # ── Discover categories that actually have data ──────────────────────────
    # Start from the categories collection (seeded on startup from ALL_CATEGORIES)
    cat_names: list[str] = []
    try:
        cursor = db[CATEGORIES_COLLECTION].find({}, {"name": 1, "_id": 0})
        docs = await cursor.to_list(length=500)
        cat_names = [d["name"] for d in docs if d.get("name")]
    except Exception:
        pass
    if not cat_names:
        cat_names = list(ALL_CATEGORIES)

    # Also consider the legacy root collection
    all_collections: list[tuple[str, str]] = []  # (display_name, collection_name)
    for cat in sorted(cat_names):
        coll_name = collection_for_category(cat)
        count = await db[coll_name].count_documents({})
        if count > 0:
            all_collections.append((cat, coll_name))

    # Legacy root collection (leads without a category slug)
    root_count = await db[COLLECTION_NAME].count_documents({})
    if root_count > 0:
        all_collections.append(("Legacy", COLLECTION_NAME))

    columns = [
        "Name", "Email", "Phone", "Company", "Designation", "Category",
        "Platform", "Form", "Status", "Notes", "Follow-up Date", "Created Date",
    ]

    # ── Shared styling helpers ───────────────────────────────────────────────
    header_fill   = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_wrap      = Alignment(wrap_text=True, vertical="top")

    def _write_sheet(ws, rows: list[dict]) -> None:
        """Write header + data rows into a worksheet and auto-fit columns."""
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                val  = row.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = top_wrap

        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = max(
                len(col_name),
                *(len(str(ws.cell(row=r, column=col_idx).value or ""))
                  for r in range(2, len(rows) + 2))
            ) if rows else len(col_name)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        if rows:
            ws.freeze_panes = "A2"

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: All Leads (merged)
    ws_all = wb.active
    ws_all.title = "All Leads"

    all_rows: list[dict] = []
    cat_rows: dict[str, list[dict]] = {}

    for display_name, coll_name in all_collections:
        cursor = db[coll_name].find({}).sort("created_at", -1).limit(10000)
        docs = await cursor.to_list(length=10000)
        rows: list[dict] = []
        for doc in docs:
            doc["id"] = str(doc.pop("_id", ""))
            rows.append(_doc_to_export_row(doc))
        cat_rows[display_name] = rows
        all_rows.extend(rows)

    _write_sheet(ws_all, all_rows)

    # One sheet per category
    for display_name, _ in all_collections:
        rows = cat_rows.get(display_name, [])
        if not rows:
            continue
        # Truncate sheet name to 31 chars (Excel limit)
        sheet_title = display_name[:31]
        ws_cat = wb.create_sheet(title=sheet_title)
        _write_sheet(ws_cat, rows)

    # ── Stream response ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"all_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /leads/origami-stats  — MUST be before /leads/{lead_id}
# Returns real Origami coverage stats calculated from the actual database.
# NEVER hardcodes percentages.
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/leads/origami-stats",
    summary="Origami enrichment coverage stats from actual database data",
    response_model=dict,
    tags=["Leads"],
)
async def get_origami_stats(
    category: Optional[str] = Query(None, description="Industry category — routes to correct collection"),
):
    """
    Calculate Origami coverage statistics from the actual MongoDB data.

    Returns:
      total_leads          — total documents in collection
      origami_enriched     — count with origami_enriched=True
      founder_found        — count with founder_name set
      founder_email_found  — count with founder_email OR email set (where origami_enriched=True)
      origami_percent      — origami_enriched / total * 100
      founder_percent      — founder_found / total * 100
      founder_email_percent — founder_email_found / founder_found * 100
      status_breakdown     — { found: N, found_decision_maker: N, not_found: N, skipped: N, error: N }
    """
    db        = get_db()
    coll_name = collection_for_category(category) if category else COLLECTION_NAME
    coll      = db[coll_name]

    total = await coll.count_documents({})
    if total == 0:
        return {
            "success": True,
            "total_leads": 0,
            "origami_enriched": 0,
            "founder_found": 0,
            "founder_email_found": 0,
            "origami_percent": 0.0,
            "founder_percent": 0.0,
            "founder_email_percent": 0.0,
            "status_breakdown": {},
        }

    origami_enriched    = await coll.count_documents({"origami_enriched": True})
    founder_found       = await coll.count_documents({"founder_name": {"$nin": [None, ""]}})
    # Email coverage: leads where origami_enriched AND (founder_email OR email) is set
    founder_email_found = await coll.count_documents({
        "origami_enriched": True,
        "$or": [
            {"founder_email": {"$nin": [None, ""]}},
            {"email":         {"$nin": [None, ""]}},
        ],
    })

    # Status breakdown
    for status_val in ("found", "found_decision_maker", "not_found", "skipped", "error"):
        pass  # Will aggregate below

    pipeline = [
        {"$group": {"_id": "$founder_status", "count": {"$sum": 1}}},
    ]
    status_cursor = coll.aggregate(pipeline)
    status_breakdown: dict = {}
    async for row in status_cursor:
        k = row.get("_id") or "unknown"
        status_breakdown[k] = row.get("count", 0)

    def _pct(num: int, denom: int) -> float:
        if denom == 0:
            return 0.0
        return round(num / denom * 100, 1)

    return {
        "success":               True,
        "total_leads":           total,
        "origami_enriched":      origami_enriched,
        "founder_found":         founder_found,
        "founder_email_found":   founder_email_found,
        "origami_percent":       _pct(origami_enriched, total),
        "founder_percent":       _pct(founder_found, total),
        "founder_email_percent": _pct(founder_email_found, max(founder_found, 1)),
        "status_breakdown":      status_breakdown,
    }


# ─────────────────────────────────────────────────────────────────────────────
# POST /leads/{lead_id}/enrich-origami
# Trigger Origami enrichment for a single existing lead.
# Does NOT create a new lead. Updates the existing document in-place.
# Falls back gracefully if Origami API key is not set.
# ─────────────────────────────────────────────────────────────────────────────

class _OrigamiEnrichBody(_StatusBodyBase):
    category: Optional[str] = None


@router.post(
    "/leads/{lead_id}/enrich-origami",
    summary="Run Origami enrichment for a single lead",
    response_model=dict,
    tags=["Leads"],
)
async def enrich_lead_with_origami(lead_id: str, payload: _OrigamiEnrichBody):
    """
    Trigger Origami enrichment for one existing lead document.

    Pipeline:
      1. Fetch lead from MongoDB
      2. Run enrich_company_with_origami()
      3. For name-only contacts, forward to Prospeo/Hunter (_enrich_origami_founder_emails)
      4. Run people waterfall (PDL → Prospeo → ContactOut → Hunter) with Origami seed
      5. Merge & dedup contacts
      6. Save back to MongoDB (never creates a duplicate)
      7. Return the updated lead document

    Fallback: if ORIGAMI_API_KEY is not set, skips Origami but still runs
    the existing people waterfall so the caller gets fresh contact data.

    Does NOT block lead generation — all failures are caught and logged.
    """
    from bson import ObjectId
    from bson.errors import InvalidId

    db        = get_db()
    now       = datetime.now(timezone.utc)
    coll_name = collection_for_category(payload.category) if payload.category else None

    try:
        oid = ObjectId(lead_id)
    except (InvalidId, Exception):
        oid = None

    # ── Find the lead document ────────────────────────────────────────────────
    doc = None

    async def _find_in(cname: str):
        coll = db[cname]
        flt  = {"_id": oid} if oid else {"id": lead_id}
        return await coll.find_one(flt)

    if coll_name:
        doc = await _find_in(coll_name)
    else:
        doc = await _find_in(COLLECTION_NAME)
        if not doc:
            cats_cursor = db[CATEGORIES_COLLECTION].find({}, {"collection": 1, "_id": 0})
            async for cat in cats_cursor:
                c = cat.get("collection")
                if c and c != COLLECTION_NAME:
                    doc = await _find_in(c)
                    if doc:
                        coll_name = c
                        break

    if not doc:
        raise HTTPException(status_code=404, detail=f"Lead '{lead_id}' not found.")

    # Determine which collection this doc lives in
    if not coll_name:
        coll_name = collection_for_category(doc.get("category") or "") or COLLECTION_NAME

    # ── Convert bson ObjectId to string so enrichment service works ───────────
    company = dict(doc)
    company["_id"] = str(company.get("_id", ""))

    t0 = time.monotonic()

    # ── Step 1: Origami enrichment (graceful skip if key not set) ─────────────
    origami_result: dict = {}
    try:
        from app.services.origami_service import (
            enrich_company_with_origami,
            is_configured as origami_configured,
            _enrich_origami_founder_emails,
        )

        if origami_configured():
            company = await enrich_company_with_origami(company)
            origami_result = {
                "origami_enriched":   company.get("origami_enriched", False),
                "origami_confidence": company.get("origami_confidence", 0.0),
                "founder_status":     company.get("founder_status", "skipped"),
                "people_count":       len(company.get("people") or []),
            }
            # Email forwarding for name-only contacts
            no_email_contacts = [
                c for c in (company.get("_origami_contacts") or [])
                if c.get("name") and not c.get("email") and company.get("domain")
            ]
            if no_email_contacts:
                company = await _enrich_origami_founder_emails(company)
        else:
            company["origami_enriched"] = False
            company["founder_status"]   = "skipped"
            origami_result = {"origami_enriched": False, "founder_status": "skipped"}

    except Exception as _orig_exc:
        print(f"[ORIGAMI] Enrichment error for {lead_id}: {_orig_exc} — continuing with waterfall")
        company.setdefault("origami_enriched", False)
        company.setdefault("founder_status", "error")
        origami_result = {"origami_enriched": False, "founder_status": "error", "error": str(_orig_exc)}

    # ── Step 2: People waterfall (PDL → Prospeo → ContactOut → Hunter) ────────
    origami_contacts = list(company.pop("_origami_contacts", None) or [])
    waterfall_result: dict = {}
    try:
        from people_enrichment.orchestrator import enrich_company_contacts, reset_cache
        reset_cache()

        company_name = company.get("company_name", "")
        domain       = company.get("domain") or ""
        website      = company.get("website") or ""

        result = await enrich_company_contacts(
            company_name=company_name,
            domain=domain or None,
            website=website or None,
            origami_contacts=origami_contacts if origami_contacts else None,
        )

        contacts_list = [
            {
                "name":         c.name,
                "title":        c.title,
                "email":        c.email,
                "phone":        c.phone,
                "linkedin_url": c.linkedin_url,
                "sources":      list(c.sources),
                "confidence":   c.confidence,
            }
            for c in result.contacts
        ]
        company["contacts"] = contacts_list

        # Promote best contact email to company level if missing
        if not company.get("email"):
            for ct in contacts_list:
                if ct.get("email"):
                    company["email"] = ct["email"]
                    break

        waterfall_result = {
            "contacts_found": result.contacts_found,
            "emails_found":   result.emails_found,
            "phones_found":   result.phones_found,
            "providers_used": result.providers_used,
        }

    except Exception as _wf_exc:
        print(f"[ORIGAMI] Waterfall error for {lead_id}: {_wf_exc}")
        waterfall_result = {"error": str(_wf_exc)}

    elapsed = round(time.monotonic() - t0, 2)

    # ── Step 3: Write back to MongoDB ─────────────────────────────────────────
    coll = db[coll_name]
    flt  = {"_id": oid} if oid else {"id": lead_id}

    set_fields = {
        "origami_enriched":    company.get("origami_enriched", False),
        "origami_confidence":  company.get("origami_confidence", 0.0),
        "origami_source":      company.get("origami_source", "origami"),
        "founder_status":      company.get("founder_status", "skipped"),
        "founder_title":       company.get("founder_title"),
        "founder_email":       company.get("founder_email"),
        "founder_profile_url": company.get("founder_profile_url"),
        "people":              company.get("people", []),
        "contacts":            company.get("contacts", []),
        "updated_at":          now,
    }
    # Promote origami-found founder fields without overwriting existing values
    if company.get("founder_name") and not doc.get("founder_name"):
        set_fields["founder_name"] = company["founder_name"]
    if company.get("founder_number") and not doc.get("founder_number"):
        set_fields["founder_number"] = company["founder_number"]
    if company.get("email") and not doc.get("email"):
        set_fields["email"] = company["email"]

    try:
        await coll.update_one(flt, {"$set": set_fields})
    except Exception as _save_exc:
        print(f"[ORIGAMI] MongoDB save error for {lead_id}: {_save_exc}")

    # ── Return the updated document ───────────────────────────────────────────
    updated_doc = await coll.find_one(flt)
    if updated_doc:
        updated_doc["id"] = str(updated_doc.pop("_id"))
        for ts in ("created_at", "updated_at"):
            if isinstance(updated_doc.get(ts), datetime):
                updated_doc[ts] = updated_doc[ts].isoformat()

    return {
        "success":         True,
        "lead_id":         lead_id,
        "elapsed_seconds": elapsed,
        "origami":         origami_result,
        "waterfall":       waterfall_result,
        "lead":            updated_doc,
    }


# ─────────────────────────────────────────────────────────────────────────────
# POST /leads/bulk-enrich-origami
# Bulk Origami enrichment for a list of lead IDs.
# Runs enrichment concurrently (max 3 at a time) without blocking generation.
# ─────────────────────────────────────────────────────────────────────────────

class _BulkOrigamiBody(_StatusBodyBase):
    lead_ids: list[str]
    category: Optional[str] = None
    max_concurrency: int = 3


@router.post(
    "/leads/bulk-enrich-origami",
    summary="Bulk Origami enrichment for a list of lead IDs",
    response_model=dict,
    tags=["Leads"],
)
async def bulk_enrich_origami(payload: _BulkOrigamiBody):
    """
    Enrich multiple leads with Origami concurrently.

    - Runs up to max_concurrency enrichments in parallel.
    - Each failure is isolated — one failure does NOT block others.
    - Returns a summary with per-lead results.
    """
    import asyncio as _asyncio

    lead_ids = payload.lead_ids or []
    if not lead_ids:
        raise HTTPException(status_code=422, detail="lead_ids cannot be empty.")
    if len(lead_ids) > 100:
        raise HTTPException(status_code=422, detail="Maximum 100 leads per bulk request.")

    sem = _asyncio.Semaphore(min(payload.max_concurrency, 5))
    results: list[dict] = []
    t0 = time.monotonic()

    async def _enrich_one(lid: str):
        async with sem:
            try:
                body = _OrigamiEnrichBody(category=payload.category)
                res  = await enrich_lead_with_origami(lid, body)
                return {"lead_id": lid, "success": True, **res.get("origami", {})}
            except HTTPException as he:
                return {"lead_id": lid, "success": False, "error": he.detail}
            except Exception as exc:
                return {"lead_id": lid, "success": False, "error": str(exc)}

    tasks = [_enrich_one(lid) for lid in lead_ids]
    results = list(await _asyncio.gather(*tasks, return_exceptions=False))

    succeeded     = sum(1 for r in results if r.get("success"))
    enriched      = sum(1 for r in results if r.get("origami_enriched"))
    founders_found= sum(1 for r in results if r.get("founder_status") in ("found", "found_decision_maker"))

    return {
        "success":        True,
        "total":          len(lead_ids),
        "succeeded":      succeeded,
        "failed":         len(lead_ids) - succeeded,
        "origami_enriched": enriched,
        "founders_found": founders_found,
        "elapsed_seconds": round(time.monotonic() - t0, 1),
        "results":        results,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Debug endpoints
# ─────────────────────────────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════════════════
# ── REDDIT SECTION  ─ POST /leads/generate-reddit ──────────────────────────
# To remove the Reddit module:
#   1. Delete this section (up to the "end REDDIT SECTION" comment)
#   2. Delete backend/reddit/
#   3. Remove "from reddit.routes import router as reddit_router" from app/main.py
#   4. Remove the Reddit source selector from the frontend
# ═══════════════════════════════════════════════════════════════════════════════

class _RedditGenerateBody(_StatusBodyBase):
    """Request payload for POST /leads/generate-reddit."""
    category: str
    location: str
    limit:    int = 25


@router.post(
    "/leads/generate-reddit",
    summary="Generate leads from Reddit posts and save to MongoDB",
    response_model=dict,
    tags=["Leads"],
)
async def generate_reddit_leads(payload: _RedditGenerateBody):
    """
    Full Reddit lead-generation pipeline:

      1. Authenticate with Reddit OAuth2 (application-only — no user login)
      2. Generate N search queries from category + location
      3. Fan-out search across all queries
      4. Extract lead candidates from relevant posts
      5. Deduplicate against existing leads in MongoDB (by post_id, email, company_name)
      6. Upsert new leads into the existing leads_{category_slug} collection
      7. Record a generation_history run (source="reddit") so History panel shows it
      8. Return summary + inserted leads

    Google Maps pipeline is completely unaffected — this runs independently.

    Error handling:
      - Reddit auth failure  → returns 200 with success=False + error detail
      - Rate limited         → returns 200 with success=False + error="rate_limited"
      - Timeout              → returns 200 with success=False + error="timeout"
      A Reddit failure never raises 5xx that would break the Google Maps pipeline.
    """
    import uuid as _uuid_mod

    # ── Guard: credentials ────────────────────────────────────────────────────
    try:
        from reddit.config import is_configured as reddit_configured
        from reddit.search import run_reddit_search, candidate_to_lead_doc
    except ImportError as exc:
        raise HTTPException(
            status_code=503,
            detail=f"Reddit module not available: {exc}",
        )

    if not reddit_configured():
        return {
            "success": False,
            "error": "no_credentials",
            "message": "REDDIT_CLIENT_ID and REDDIT_CLIENT_SECRET are not set in .env",
            "category": payload.category,
            "location": payload.location,
        }

    t_start = time.monotonic()
    category = payload.category.strip()
    location = payload.location.strip()
    limit    = max(1, min(payload.limit, 100))

    # Generate a unique run ID for the history record
    run_id = "RUN-RDT-" + _uuid_mod.uuid4().hex[:8].upper()

    _log("REDDIT", f"generate-reddit started — category={category!r} location={location!r} limit={limit} run_id={run_id}")

    # ── Create generation_history run (source="reddit") ───────────────────────
    db = get_db()
    now_utc = datetime.now(timezone.utc)
    history_doc = {
        "run_id":          run_id,
        "category":        category,
        "search_query":    f"{category} {location}",
        "state":           location,
        "district":        "",
        "requested_count": limit,
        "generated_count": 0,
        "updated_count":   0,
        "status":          "running",
        "source":          "reddit",
        "started_at":      now_utc.isoformat(),
        "created_at":      now_utc.isoformat(),
        "completed_at":    None,
        "failed_at":       None,
        "duration_seconds": None,
        "filters": {"category": category, "location": location, "limit": limit},
        "lead_ids":        [],
        "logs": [
            {
                "timestamp": datetime.now().strftime("%H:%M:%S"),
                "level": "INFO",
                "stage": "init",
                "message": f"Reddit generation started — category={category!r} location={location!r} limit={limit}",
            }
        ],
        "statistics":      {},
        "error_message":   None,
        "pipeline_stats":  None,
    }
    try:
        await db[HISTORY_COLLECTION].insert_one(history_doc)
    except Exception as hist_exc:
        _log("REDDIT", f"WARNING: could not create history run: {hist_exc}")

    async def _append_reddit_log(level: str, stage: str, message: str) -> None:
        entry = {"timestamp": datetime.now().strftime("%H:%M:%S"),
                 "level": level, "stage": stage, "message": message}
        try:
            await db[HISTORY_COLLECTION].update_one(
                {"run_id": run_id}, {"$push": {"logs": entry}}
            )
        except Exception:
            pass

    # ── Run Reddit search ──────────────────────────────────────────────────────
    await _append_reddit_log("SEARCH", "reddit", f"Searching Reddit for {category!r} in {location!r}")

    try:
        search_result = await run_reddit_search(
            category=category,
            location=location,
            limit=limit,
        )
    except Exception as exc:
        error_msg = str(exc)
        _log("REDDIT", f"Search error: {error_msg}")
        await _append_reddit_log("ERROR", "reddit", f"Search failed: {error_msg}")
        now_f = datetime.now(timezone.utc)
        await db[HISTORY_COLLECTION].update_one(
            {"run_id": run_id},
            {"$set": {
                "status": "failed",
                "failed_at": now_f.isoformat(),
                "duration_seconds": round(time.monotonic() - t_start, 1),
                "error_message": error_msg,
            }},
        )
        return {
            "success": False,
            "run_id": run_id,
            "category": category,
            "location": location,
            "error": error_msg,
            "total_discovered": 0,
            "total_valid": 0,
            "total_inserted": 0,
            "total_duplicates": 0,
            "total_failed": 0,
            "leads": [],
        }

    candidates = search_result.get("candidates", [])
    posts_discovered = search_result.get("posts_discovered", 0)
    queries_run = search_result.get("queries_run", 0)
    search_error = search_result.get("error")

    _log("REDDIT", f"Posts discovered: {posts_discovered} | Valid candidates: {len(candidates)}")
    await _append_reddit_log("SEARCH", "reddit",
        f"Posts discovered: {posts_discovered} | Valid candidates: {len(candidates)}")

    if search_error and not candidates:
        now_f = datetime.now(timezone.utc)
        await db[HISTORY_COLLECTION].update_one(
            {"run_id": run_id},
            {"$set": {
                "status": "failed",
                "failed_at": now_f.isoformat(),
                "duration_seconds": round(time.monotonic() - t_start, 1),
                "error_message": search_error,
            }},
        )
        return {
            "success": False,
            "run_id": run_id,
            "category": category,
            "location": location,
            "error": search_error,
            "total_discovered": posts_discovered,
            "total_valid": 0,
            "total_inserted": 0,
            "total_duplicates": 0,
            "total_failed": 0,
            "leads": [],
        }

    # ── MongoDB deduplication & upsert ────────────────────────────────────────
    coll_name = collection_for_category(category)
    coll = db[coll_name]

    # Ensure indexes (non-fatal)
    try:
        await ensure_lead_indexes(db, category)
    except Exception:
        pass

    # Register category
    try:
        await db[CATEGORIES_COLLECTION].update_one(
            {"name": category},
            {"$set": {"name": category, "collection": coll_name}},
            upsert=True,
        )
    except Exception:
        pass

    # Collect dedup keys from candidates
    candidate_post_ids    = [c.post_id for c in candidates if c.post_id]
    candidate_emails      = [c.email for c in candidates if c.email]
    candidate_companies   = [
        (c.company_name or "").lower().strip()
        for c in candidates if c.company_name
    ]

    # Existing post_ids (Reddit-specific dedup)
    existing_post_ids: set[str] = set()
    if candidate_post_ids:
        async for doc in coll.find(
            {"post_id": {"$in": candidate_post_ids}},
            {"post_id": 1, "_id": 0},
        ):
            pid = doc.get("post_id")
            if pid:
                existing_post_ids.add(pid)

    # Existing emails
    existing_emails: set[str] = set()
    if candidate_emails:
        async for doc in coll.find(
            {"email": {"$in": candidate_emails}},
            {"email": 1, "_id": 0},
        ):
            e = (doc.get("email") or "").lower().strip()
            if e:
                existing_emails.add(e)

    # Existing company names (only when no email/post_id match)
    existing_companies: set[str] = set()
    if candidate_companies:
        async for doc in coll.find(
            {"company_name": {"$regex": "|".join(
                c for c in candidate_companies if c
            ), "$options": "i"}} if candidate_companies else {},
            {"company_name": 1, "_id": 0},
        ):
            n = (doc.get("company_name") or "").lower().strip()
            if n:
                existing_companies.add(n)

    _log("REDDIT",
         f"DB dedup check — existing post_ids={len(existing_post_ids)} "
         f"emails={len(existing_emails)} companies={len(existing_companies)}")
    await _append_reddit_log("FILTER", "dedup",
        f"DB dedup: {len(existing_post_ids)} existing post IDs, "
        f"{len(existing_emails)} existing emails")

    # ── Upsert loop ───────────────────────────────────────────────────────────
    inserted_count  = 0
    duplicate_count = 0
    failed_count    = 0
    inserted_ids: list[str] = []
    now_ins = datetime.now(timezone.utc)

    for candidate in candidates:
        # Check duplicates
        post_dup    = candidate.post_id in existing_post_ids
        email_dup   = bool(candidate.email and candidate.email.lower() in existing_emails)
        company_dup = bool(
            candidate.company_name
            and candidate.company_name.lower().strip() in existing_companies
            and not candidate.email  # only name-dedup when no email
        )

        if post_dup or email_dup or company_dup:
            duplicate_count += 1
            continue

        # Build lead document
        try:
            lead_doc = candidate_to_lead_doc(candidate, category, run_id)
        except Exception as conv_exc:
            _log("REDDIT", f"Conversion error for post {candidate.post_id}: {conv_exc}")
            failed_count += 1
            continue

        # Upsert key: post_id is the most reliable Reddit dedup key
        filter_key: dict
        if candidate.post_id:
            filter_key = {"post_id": candidate.post_id}
        elif candidate.email:
            filter_key = {"email": candidate.email}
        else:
            filter_key = {"company_name": lead_doc["company_name"]}

        set_on_insert = {
            "created_at":          now_ins,
            "generation_run_id":   run_id,
            "status":              "new",
            "status_updated_at":   now_ins.isoformat(),
            "notes":               [],
            "follow_up_date":      None,
            "contacts":            [],
            "people":              [],
            "origami_enriched":    False,
        }

        try:
            result = await coll.update_one(
                filter_key,
                {
                    "$set": lead_doc,
                    "$setOnInsert": set_on_insert,
                },
                upsert=True,
            )
            if result.upserted_id:
                inserted_count += 1
                inserted_ids.append(str(result.upserted_id))
                # Track for next iteration dedup
                if candidate.post_id:
                    existing_post_ids.add(candidate.post_id)
                if candidate.email:
                    existing_emails.add(candidate.email.lower())
            else:
                duplicate_count += 1  # already existed (race or filter_key matched)
        except Exception as db_exc:
            _log("REDDIT", f"DB upsert error for post {candidate.post_id}: {db_exc}")
            failed_count += 1

    elapsed_upsert = round(time.monotonic() - t_start, 1)
    _log("REDDIT",
         f"Upserts complete — inserted={inserted_count} duplicates={duplicate_count} "
         f"failed={failed_count} elapsed={elapsed_upsert}s")
    await _append_reddit_log("COMPLETE", "pipeline",
        f"Upserts done in {elapsed_upsert}s — "
        f"inserted={inserted_count} duplicates={duplicate_count} failed={failed_count}")

    # ── Fetch inserted leads from MongoDB ─────────────────────────────────────
    leads_out: list[dict] = []
    raw_oids: list = []
    if inserted_ids:
        from bson import ObjectId
        from bson.errors import InvalidId
        for sid in inserted_ids:
            try:
                raw_oids.append(ObjectId(sid))
            except (InvalidId, Exception):
                pass
        if raw_oids:
            cursor = coll.find({"_id": {"$in": raw_oids}}).sort("created_at", -1)
            db_docs = await cursor.to_list(length=len(raw_oids) + 5)
            for doc in db_docs:
                doc["id"] = str(doc.pop("_id"))
                for ts_f in ("created_at", "updated_at"):
                    if isinstance(doc.get(ts_f), datetime):
                        doc[ts_f] = doc[ts_f].isoformat()
                leads_out.append(doc)

    # ── People enrichment waterfall (PDL → Prospeo → ContactOut → Hunter) ────
    # Reuses the EXACT same orchestrator used by the Google Maps pipeline.
    # Runs only on newly inserted leads; enrichment failures are non-fatal.
    enrichment_stats: dict = {}
    if leads_out:
        await _append_reddit_log("ENRICH", "people_enrichment",
            f"Starting people enrichment for {len(leads_out)} new Reddit leads")
        _log("REDDIT", f"Starting people enrichment for {len(leads_out)} leads")
        try:
            from reddit.enrichment import enrich_reddit_leads_batch
            leads_out, enrichment_stats = await enrich_reddit_leads_batch(
                leads_out,
                max_concurrency=3,
                per_lead_timeout=60.0,
                log=lambda msg: _log("REDDIT_ENRICH", msg),
            )
            _log("REDDIT", (
                f"Enrichment complete — "
                f"enriched={enrichment_stats.get('enriched', 0)} "
                f"contacts={enrichment_stats.get('contacts_found', 0)} "
                f"emails={enrichment_stats.get('emails_found', 0)} "
                f"elapsed={enrichment_stats.get('elapsed_seconds', 0)}s"
            ))
            await _append_reddit_log("ENRICH", "people_enrichment",
                f"Enrichment done — "
                f"enriched={enrichment_stats.get('enriched', 0)} "
                f"contacts={enrichment_stats.get('contacts_found', 0)} "
                f"emails={enrichment_stats.get('emails_found', 0)}")
        except Exception as enrich_exc:
            # Enrichment failure is completely non-fatal — leads are still returned
            _log("REDDIT", f"WARNING: enrichment step error (non-fatal): {enrich_exc}")
            await _append_reddit_log("ENRICH", "people_enrichment",
                f"Enrichment step error (non-fatal): {enrich_exc}")

        # ── Write enrichment results back to MongoDB ──────────────────────────
        # Only update fields that enrichment actually filled in.
        # Preserves all existing Reddit source fields.
        if leads_out:
            now_enrich = datetime.now(timezone.utc)
            for enriched_lead in leads_out:
                lead_id_str = enriched_lead.get("id") or ""
                if not lead_id_str:
                    continue
                try:
                    from bson import ObjectId as _ObjId
                    from bson.errors import InvalidId as _InvId
                    lead_oid = _ObjId(lead_id_str)
                except Exception:
                    continue

                # Build $set with only non-empty enrichment fields
                enrich_set: dict = {"updated_at": now_enrich}

                contacts = enriched_lead.get("contacts") or []
                if contacts:
                    enrich_set["contacts"] = contacts

                # Only promote email when enrichment added one
                enr_email = enriched_lead.get("email")
                if enr_email:
                    enrich_set["email"] = enr_email
                enr_emails = enriched_lead.get("emails") or []
                if enr_emails:
                    enrich_set["emails"] = enr_emails

                enr_phone = enriched_lead.get("company_number")
                if enr_phone:
                    enrich_set["company_number"] = enr_phone
                enr_phones = enriched_lead.get("phones") or []
                if enr_phones:
                    enrich_set["phones"] = enr_phones

                enr_founder = enriched_lead.get("founder_name")
                if enr_founder:
                    enrich_set["founder_name"] = enr_founder
                enr_founder_num = enriched_lead.get("founder_number")
                if enr_founder_num:
                    enrich_set["founder_number"] = enr_founder_num

                enr_conf = enriched_lead.get("confidence")
                if enr_conf is not None:
                    enrich_set["confidence"] = enr_conf

                enr_fv = enriched_lead.get("_field_verification")
                if enr_fv:
                    enrich_set["_field_verification"] = enr_fv

                pe_stats = enriched_lead.get("people_enrichment_stats")
                if pe_stats:
                    enrich_set["people_enrichment_stats"] = pe_stats

                # Ensure Reddit source fields are never wiped
                enrich_set["source"]          = "reddit"
                enrich_set["platform"]        = "reddit"
                enrich_set["research_source"] = "reddit"

                try:
                    await coll.update_one(
                        {"_id": lead_oid},
                        {"$set": enrich_set},
                    )
                except Exception as upd_exc:
                    _log("REDDIT", f"WARNING: enrichment write-back failed for {lead_id_str}: {upd_exc}")

    # Update lead_count in categories
    try:
        total_in_coll = await coll.count_documents({})
        await db[CATEGORIES_COLLECTION].update_one(
            {"name": category},
            {"$set": {"lead_count": total_in_coll}},
        )
    except Exception:
        pass

    # ── Complete the history run ───────────────────────────────────────────────
    elapsed = round(time.monotonic() - t_start, 1)
    run_stats = {
        "posts_discovered":  posts_discovered,
        "candidates_found":  len(candidates),
        "leads_generated":   inserted_count,
        "duplicates":        duplicate_count,
        "failed":            failed_count,
        "queries_run":       queries_run,
        "elapsed_seconds":   elapsed,
        # Enrichment stats
        "enriched":          enrichment_stats.get("enriched", 0),
        "contacts_found":    enrichment_stats.get("contacts_found", 0),
        "emails_found":      enrichment_stats.get("emails_found", 0),
    }
    now_done = datetime.now(timezone.utc)
    try:
        await db[HISTORY_COLLECTION].update_one(
            {"run_id": run_id},
            {"$set": {
                "status":          "completed",
                "generated_count": inserted_count,
                "updated_count":   duplicate_count,
                "completed_at":    now_done.isoformat(),
                "duration_seconds": elapsed,
                "lead_ids":        inserted_ids,
                "statistics":      run_stats,
                "pipeline_stats":  {
                    "reddit_posts_discovered": posts_discovered,
                    "reddit_queries_run":      queries_run,
                    "reddit_candidates":       len(candidates),
                    "elapsed_seconds":         elapsed,
                    # People enrichment stats
                    "people_enriched":         enrichment_stats.get("enriched", 0),
                    "people_contacts_found":   enrichment_stats.get("contacts_found", 0),
                    "people_emails_found":     enrichment_stats.get("emails_found", 0),
                    "people_elapsed_seconds":  enrichment_stats.get("elapsed_seconds", 0.0),
                },
            }},
        )
    except Exception as hist_upd_exc:
        _log("REDDIT", f"WARNING: could not complete history run: {hist_upd_exc}")

    return {
        "success":          True,
        "run_id":           run_id,
        "category":         category,
        "location":         location,
        "total_discovered": posts_discovered,
        "total_valid":      len(candidates),
        "total_inserted":   inserted_count,
        "total_duplicates": duplicate_count,
        "total_failed":     failed_count,
        "elapsed_seconds":  elapsed,
        "leads":            leads_out,
        "pipeline_stats": {
            "reddit_posts_discovered": posts_discovered,
            "reddit_queries_run":      queries_run,
            "reddit_candidates":       len(candidates),
            "elapsed_seconds":         elapsed,
            # People enrichment stats
            "people_enriched":         enrichment_stats.get("enriched", 0),
            "people_contacts_found":   enrichment_stats.get("contacts_found", 0),
            "people_emails_found":     enrichment_stats.get("emails_found", 0),
            "people_elapsed_seconds":  enrichment_stats.get("elapsed_seconds", 0.0),
        },
    }


# ── end REDDIT SECTION ────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────────
# Debug endpoints
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/debug/database",
    summary="MongoDB connectivity check",
    response_model=dict,
    tags=["Debug"],
)
async def debug_database():
    """Show connection status, database name, and per-category lead counts."""
    db   = get_db()
    coll = db[COLLECTION_NAME]

    # Legacy leads count
    legacy_count = await coll.count_documents({})

    # Per-category counts
    cats_coll   = db[CATEGORIES_COLLECTION]
    cats_cursor = cats_coll.find({}, {"name": 1, "collection": 1, "_id": 0}).sort("name", 1)
    cats_docs   = await cats_cursor.to_list(length=500)

    category_counts: dict = {}
    for cat in cats_docs:
        cat_name  = cat.get("name", "")
        cat_coll  = cat.get("collection") or collection_for_category(cat_name)
        count     = await db[cat_coll].count_documents({})
        if count > 0:
            category_counts[cat_name] = count

    return {
        "connected":       True,
        "database":        db.name,
        "legacy_collection": COLLECTION_NAME,
        "legacy_count":    legacy_count,
        "categories_collection": CATEGORIES_COLLECTION,
        "category_lead_counts":  category_counts,
    }


@router.get(
    "/debug/sample",
    summary="First 5 documents from leads collection",
    response_model=list[dict[str, Any]],
    tags=["Debug"],
)
async def debug_sample():
    db     = get_db()
    coll   = db[COLLECTION_NAME]
    cursor = coll.find({}, {"_id": 0}).limit(5)
    docs   = await cursor.to_list(length=5)
    return docs
